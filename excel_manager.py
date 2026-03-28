"""
Excel Database Manager for Smart Parking System
Handles all CRUD operations on the Excel-based database.
"""

import os
import threading
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash
import config

def get_ist_now():
    """Get current time in IST (UTC+5:30)."""
    return datetime.utcnow() + timedelta(hours=5, minutes=30)

# Thread safe re-entrant lock to prevent deadlocks
excel_lock = threading.RLock()


def init_excel():
    """Initialize or migrate the Excel database with required sheets and headers."""
    required_sheets = {
        'Users': ['UserID', 'Name', 'Email', 'Password', 'Phone', 'Role', 'PlateNumber', 'PapersUrl', 'LicenseUrl', 'LastActive'],
        'ParkingSlots': ['SlotID', 'SlotNumber', 'Status'],
        'Bookings': ['BookingID', 'UserID', 'SlotID', 'SlotNumber', 'Date', 'Time', 'UserName', 'UserEmail', 'UserStatus', 'LoginTime', 'LogoutTime'],
        'ActivityLogs': ['LogID', 'UserID', 'UserName', 'UserEmail', 'Action', 'Date', 'Time'],
        'Feedbacks': ['FeedbackID', 'Name', 'Email', 'Rating', 'Feedback', 'Date', 'Time', 'CreatedAt']
    }

    if not os.path.exists(config.EXCEL_FILE):
        wb = Workbook()
        # Initialize each required sheet
        for i, (title, headers) in enumerate(required_sheets.items()):
            if i == 0:
                ws = wb.active
                ws.title = title
            else:
                ws = wb.create_sheet(title)
            ws.append(headers)

            # Default slots
            if title == 'ParkingSlots':
                for j in range(1, config.TOTAL_SLOTS + 1):
                    ws.append([j, f'P-{j:03d}', 'Available'])
            
            # Default users (demo)
            if title == 'Users':
                ws.append([1, 'Test User', 'test@example.com', generate_password_hash('123456'), '1234567890', 'User', 'N/A', 'N/A', 'N/A', 'N/A'])
        
        wb.save(config.EXCEL_FILE)
        print(f"[INFO] New Excel database initialized.")
        return

    # Migration: Update existing file headers
    wb = _load_workbook()
    changed = False
    for title, headers in required_sheets.items():
        if title not in wb.sheetnames:
            ws = wb.create_sheet(title)
            ws.append(headers)
            changed = True
        else:
            ws = wb[title]
            first_row = [cell.value for cell in ws[1]]
            if first_row != headers:
                # Update header row while preserving other data
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)
                changed = True
    
    if changed:
        _save_workbook(wb)
        print("[INFO] Excel database migrated to new schema.")


def _load_workbook():
    """Load the workbook safely. Re-initializes if file is missing."""
    if not os.path.exists(config.EXCEL_FILE):
        print(f"[WARN] Excel file missing at {config.EXCEL_FILE}. Re-initializing...")
        init_excel()
    return load_workbook(config.EXCEL_FILE)


def _save_workbook(wb):
    """Save the workbook safely."""
    wb.save(config.EXCEL_FILE)


# ─── USER OPERATIONS ────────────────────────────────────────────

def get_user_by_email(email):
    """Find a user by email address."""
    with excel_lock:
        wb = _load_workbook()
        ws = wb['Users']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[2] and row[2].lower() == email.lower():
                return {
                    'UserID': row[0],
                    'Name': row[1],
                    'Email': row[2],
                    'Password': row[3],
                    'Phone': row[4] if len(row) > 4 else '',
                    'Role': row[5] if len(row) > 5 else 'User',
                    'PlateNumber': row[6] if len(row) > 6 else 'N/A',
                    'PapersUrl': row[7] if len(row) > 7 else 'N/A',
                    'LicenseUrl': row[8] if len(row) > 8 else 'N/A',
                    'LastActive': row[9] if len(row) > 9 else 'N/A'
                }
        return None


def get_user_by_id(user_id):
    """Find a user by their UserID with type-agnostic comparison."""
    with excel_lock:
        wb = _load_workbook()
        ws = wb['Users']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and str(row[0]).strip() == str(user_id).strip():
                return {
                    'UserID': row[0],
                    'Name': row[1],
                    'Email': row[2],
                    'Password': row[3],
                    'Phone': row[4] if len(row) > 4 else 'N/A',
                    'Role': row[5] if len(row) > 5 else 'User',
                    'PlateNumber': row[6] if len(row) > 6 else 'N/A',
                    'PapersUrl': row[7] if len(row) > 7 else 'N/A',
                    'LicenseUrl': row[8] if len(row) > 8 else 'N/A',
                    'LastActive': row[9] if len(row) > 9 else 'N/A'
                }
        return None


def register_user(name, email, password, phone, role='User', plate_number='N/A', papers_url='N/A', license_url='N/A'):
    """Register a new user. Returns the new user dict or None if email exists."""
    with excel_lock:
        wb = _load_workbook()
        ws = wb['Users']

        # Check if email already exists
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[2] and row[2].lower() == email.lower():
                return None

        # Generate new UserID
        max_id = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and isinstance(row[0], int):
                max_id = max(max_id, row[0])
        new_id = max_id + 1

        ws.append([new_id, name, email, password, phone, role, plate_number, papers_url, license_url, 'N/A'])
        _save_workbook(wb)

        return {
            'UserID': new_id,
            'Name': name,
            'Email': email,
            'Phone': phone,
            'Role': role,
            'PlateNumber': plate_number,
            'PapersUrl': papers_url,
            'LicenseUrl': license_url
        }


def get_all_users():
    """Get all registered users."""
    with excel_lock:
        wb = _load_workbook()
        ws = wb['Users']
        users = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                users.append({
                    'UserID': row[0],
                    'Name': row[1],
                    'Email': row[2],
                    'Phone': row[4] if len(row) > 4 else '',
                    'Role': row[5] if len(row) > 5 else 'User',
                    'PlateNumber': row[6] if len(row) > 6 else 'N/A',
                    'PapersUrl': row[7] if len(row) > 7 else 'N/A',
                    'LicenseUrl': row[8] if len(row) > 8 else 'N/A',
                    'LastActive': row[9] if len(row) > 9 else 'N/A'
                })
        return users


def update_user_activity(user_id):
    """Update the LastActive timestamp for a user."""
    now = get_ist_now().strftime("%Y-%m-%d %H:%M:%S")
    with excel_lock:
        wb = _load_workbook()
        ws = wb['Users']
        updated = False
        for row in ws.iter_rows(min_row=2):
            if row[0].value == user_id:
                # LastActive is now column 10
                ws.cell(row=row[0].row, column=10, value=now)
                updated = True
                break
        if updated:
            _save_workbook(wb)
        return updated


# ─── PARKING SLOT OPERATIONS ────────────────────────────────────

def get_all_slots():
    """Get all parking slots with their status."""
    with excel_lock:
        wb = _load_workbook()
        ws = wb['ParkingSlots']
        slots = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                slots.append({
                    'SlotID': row[0],
                    'SlotNumber': row[1],
                    'Status': row[2]
                })
        return slots


def get_slot_by_id(slot_id):
    """Get a single slot by its ID."""
    with excel_lock:
        wb = _load_workbook()
        ws = wb['ParkingSlots']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == slot_id:
                return {
                    'SlotID': row[0],
                    'SlotNumber': row[1],
                    'Status': row[2]
                }
        return None


def update_slot_status(slot_id, status):
    """Update the status of a parking slot."""
    with excel_lock:
        wb = _load_workbook()
        ws = wb['ParkingSlots']
        updated = False
        for row in ws.iter_rows(min_row=2):
            if row[0].value == slot_id:
                row[2].value = status
                updated = True
                break
        if updated:
            _save_workbook(wb)
        return updated


def add_slot(slot_number):
    """Add a new parking slot."""
    with excel_lock:
        wb = _load_workbook()
        ws = wb['ParkingSlots']

        # Check if slot number already exists
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[1] == slot_number:
                return None

        # Generate new SlotID
        max_id = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and isinstance(row[0], int):
                max_id = max(max_id, row[0])
        new_id = max_id + 1

        ws.append([new_id, slot_number, 'Available'])
        _save_workbook(wb)
        return {'SlotID': new_id, 'SlotNumber': slot_number, 'Status': 'Available'}


def delete_slot(slot_id):
    """Delete a parking slot (only if Available)."""
    with excel_lock:
        wb = _load_workbook()
        ws = wb['ParkingSlots']
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == slot_id:
                if row[2].value == 'Booked':
                    return False
                ws.delete_rows(idx)
                _save_workbook(wb)
                return True
        return False


# ─── BOOKING OPERATIONS ─────────────────────────────────────────

def create_booking(user_id, slot_id):
    """Create a new booking. Returns booking dict or None if slot unavailable."""
    with excel_lock:
        wb = _load_workbook()
        ws_slots = wb['ParkingSlots']
        ws_bookings = wb['Bookings']
        ws_users = wb['Users']

        # Find the slot and check availability
        slot_row = None
        slot_number = None
        for row in ws_slots.iter_rows(min_row=2):
            if row[0].value == slot_id:
                if row[2].value != 'Available':
                    return None  # Slot not available
                slot_row = row
                slot_number = row[1].value
                break

        if not slot_row:
            return None

        # Get user info
        user_name = ''
        user_email = ''
        for row in ws_users.iter_rows(min_row=2, values_only=True):
            if row and row[0] == user_id:
                user_name = row[1]
                user_email = row[2]
                break

        # Mark slot as booked
        slot_row[2].value = 'Booked'

        # Generate new BookingID
        max_id = 0
        for row in ws_bookings.iter_rows(min_row=2, values_only=True):
            if row and row[0] and isinstance(row[0], int):
                max_id = max(max_id, row[0])
        new_id = max_id + 1

        now = get_ist_now()
        date_str = now.strftime('%Y-%m-%d')
        time_str = now.strftime('%H:%M:%S')

        ws_bookings.append([new_id, user_id, slot_id, slot_number, date_str, time_str, user_name, user_email, 'Pending', 'N/A', 'N/A'])
        _save_workbook(wb)

        return {
            'BookingID': new_id,
            'UserID': user_id,
            'SlotID': slot_id,
            'SlotNumber': slot_number,
            'Date': date_str,
            'Time': time_str,
            'UserName': user_name,
            'UserEmail': user_email,
            'UserStatus': 'Pending',
            'LoginTime': 'N/A',
            'LogoutTime': 'N/A'
        }


def get_booking_by_id(booking_id):
    """Get booking details by its ID."""
    with excel_lock:
        wb = _load_workbook()
        ws = wb['Bookings']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == booking_id:
                return {
                    'BookingID': row[0],
                    'UserID': row[1],
                    'SlotID': row[2],
                    'SlotNumber': row[3],
                    'Date': row[4],
                    'Time': row[5],
                    'UserName': row[6],
                    'UserEmail': row[7],
                    'UserStatus': row[8] if len(row) > 8 else 'Pending',
                    'LoginTime': row[9] if len(row) > 9 else 'N/A',
                    'LogoutTime': row[10] if len(row) > 10 else 'N/A'
                }
        return None


def update_booking_access_status(booking_id, status, time=None):
    """Update user status and login/logout time in a booking."""
    with excel_lock:
        wb = _load_workbook()
        ws = wb['Bookings']
        updated = False
        for row in ws.iter_rows(min_row=2):
            if row[0].value == booking_id:
                row[8].value = status
                if status == 'Logged In' and time:
                    row[9].value = time
                elif status == 'Logged Out' and time:
                    row[10].value = time
                updated = True
                break
        if updated:
            _save_workbook(wb)
        return updated


def cancel_booking(booking_id):
    """Cancel a booking and free the slot."""
    with excel_lock:
        wb = _load_workbook()
        ws_bookings = wb['Bookings']
        ws_slots = wb['ParkingSlots']

        for idx, row in enumerate(ws_bookings.iter_rows(min_row=2), start=2):
            if row[0].value == booking_id:
                slot_id = row[2].value

                # Free the slot
                for slot_row in ws_slots.iter_rows(min_row=2):
                    if slot_row[0].value == slot_id:
                        slot_row[2].value = 'Available'
                        break

                ws_bookings.delete_rows(idx)
                _save_workbook(wb)
                return True
        return False


def get_all_bookings():
    """Get all bookings."""
    with excel_lock:
        wb = _load_workbook()
        ws = wb['Bookings']
        bookings = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                bookings.append({
                    'BookingID': row[0],
                    'UserID': row[1],
                    'SlotID': row[2],
                    'SlotNumber': row[3],
                    'Date': row[4],
                    'Time': row[5],
                    'UserName': row[6],
                    'UserEmail': row[7],
                    'UserStatus': row[8] if len(row) > 8 else 'Pending',
                    'LoginTime': row[9] if len(row) > 9 else 'N/A',
                    'LogoutTime': row[10] if len(row) > 10 else 'N/A'
                })
        return bookings


def get_user_bookings(user_id):
    """Get all bookings for a specific user."""
    with excel_lock:
        wb = _load_workbook()
        ws = wb['Bookings']
        bookings = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[1] == user_id:
                bookings.append({
                    'BookingID': row[0],
                    'UserID': row[1],
                    'SlotID': row[2],
                    'SlotNumber': row[3],
                    'Date': row[4],
                    'Time': row[5],
                    'UserName': row[6],
                    'UserEmail': row[7],
                    'UserStatus': row[8] if len(row) > 8 else 'Pending',
                    'LoginTime': row[9] if len(row) > 9 else 'N/A',
                    'LogoutTime': row[10] if len(row) > 10 else 'N/A'
                })
        return bookings


def get_dashboard_stats():
    """Get statistics for the admin dashboard."""
    with excel_lock:
        wb = _load_workbook()

        # Count users
        ws_users = wb['Users']
        total_users = sum(1 for _ in ws_users.iter_rows(min_row=2, values_only=True) if _[0])

        # Count slots
        ws_slots = wb['ParkingSlots']
        total_slots = 0
        available_slots = 0
        booked_slots = 0
        for row in ws_slots.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                total_slots += 1
                if row[2] == 'Available':
                    available_slots += 1
                else:
                    booked_slots += 1

        # Count bookings and parked vehicles
        ws_bookings = wb['Bookings']
        total_bookings = 0
        parked_vehicles = 0
        for row in ws_bookings.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                total_bookings += 1
                if len(row) > 8 and row[8] == 'Logged In':
                    parked_vehicles += 1

        return {
            'total_users': total_users,
            'total_slots': total_slots,
            'available_slots': available_slots,
            'booked_slots': booked_slots,
            'total_bookings': total_bookings,
            'parked_vehicles': parked_vehicles
        }


def get_full_dashboard_data():
    """Get all dashboard data in a single call to avoid multiple file locks/loads."""
    with excel_lock:
        wb = _load_workbook()

        # Users
        ws_users = wb['Users']
        users = []
        for row in ws_users.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                users.append({'UserID': row[0]})

        # Slots
        ws_slots = wb['ParkingSlots']
        slots = []
        available_slots = 0
        booked_slots = 0
        for row in ws_slots.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                slots.append({
                    'SlotID': row[0],
                    'SlotNumber': row[1],
                    'Status': row[2]
                })
                if row[2] == 'Available':
                    available_slots += 1
                else:
                    booked_slots += 1

        # Bookings
        ws_bookings = wb['Bookings']
        bookings = []
        parked_vehicles = 0
        for row in ws_bookings.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                status = row[8] if len(row) > 8 else 'Pending'
                bookings.append({
                    'BookingID': row[0],
                    'UserID': row[1],
                    'SlotID': row[2],
                    'SlotNumber': row[3],
                    'Date': row[4],
                    'Time': row[5],
                    'UserName': row[6],
                    'UserEmail': row[7],
                    'UserStatus': status,
                    'LoginTime': row[9] if len(row) > 9 else 'N/A',
                    'LogoutTime': row[10] if len(row) > 10 else 'N/A'
                })
                if status == 'Logged In' or status == 'Checked In':
                    parked_vehicles += 1

        return {
            'stats': {
                'total_users': len(users),
                'total_slots': len(slots),
                'available_slots': available_slots,
                'booked_slots': booked_slots,
                'total_bookings': len(bookings),
                'parked_vehicles': parked_vehicles
            },
            'slots': slots,
            'bookings': bookings
        }

# ─── ACTIVITY LOGS ─────────────────────────────────────────────

def log_activity(user_id, name, email, action):
    """Log user/admin activity."""
    with excel_lock:
        wb = _load_workbook()
        ws = wb['ActivityLogs']

        # Generate new LogID
        max_id = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and isinstance(row[0], int):
                max_id = max(max_id, row[0])
        new_id = max_id + 1

        now = get_ist_now()
        date_str = now.strftime('%Y-%m-%d')
        time_str = now.strftime('%H:%M:%S')

        ws.append([new_id, user_id, name, email, action, date_str, time_str])
        _save_workbook(wb)
        return True


def get_all_logs(limit=100):
    """Get all activity logs, newest first."""
    with excel_lock:
        wb = _load_workbook()
        ws = wb['ActivityLogs']
        logs = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                logs.append({
                    'LogID': row[0],
                    'UserID': row[1],
                    'UserName': row[2],
                    'UserEmail': row[3],
                    'Action': row[4],
                    'Date': row[5],
                    'Time': row[6]
                })
        
        # Sort by date and time descending
        logs.sort(key=lambda x: (x['Date'], x['Time']), reverse=True)
        return logs[:limit]

# ─── FEEDBACK OPERATIONS ────────────────────────────────────────

def save_feedback(name, email, rating, feedback_text):
    """Save user feedback to Excel."""
    with excel_lock:
        wb = _load_workbook()
        if 'Feedbacks' not in wb.sheetnames:
            ws = wb.create_sheet('Feedbacks')
            ws.append(['FeedbackID', 'Name', 'Email', 'Rating', 'Feedback', 'Date', 'Time', 'CreatedAt'])
        else:
            ws = wb['Feedbacks']

        # Generate new FeedbackID
        max_id = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and isinstance(row[0], int):
                max_id = max(max_id, row[0])
        new_id = max_id + 1

        now = get_ist_now()
        ws.append([new_id, name, email, rating, feedback_text, now.strftime('%Y-%m-%d'), now.strftime('%H:%M:%S'), now.isoformat()])
        _save_workbook(wb)
        return True

def get_all_feedbacks():
    """Get all user feedbacks from Excel."""
    with excel_lock:
        wb = _load_workbook()
        if 'Feedbacks' not in wb.sheetnames:
            return []
            
        ws = wb['Feedbacks']
        feedbacks = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                feedbacks.append({
                    'FeedbackID': row[0],
                    'name': row[1],
                    'email': row[2],
                    'rating': int(row[3]) if row[3] is not None else 0,
                    'feedback': row[4],
                    'Date': row[5],
                    'Time': row[6],
                    'createdAt': row[7]
                })
        
        # Sort by CreatedAt descending
        feedbacks.sort(key=lambda x: x.get('createdAt', ''), reverse=True)
        return feedbacks
